#!/usr/bin/python3

from absl import app, flags
import openpyxl
import numpy as np
import tensorflow as tf

FLAGS = flags.FLAGS

def add_options():
  flags.DEFINE_string('input', default = None, help = 'path to xls')
  flags.DEFINE_string('output', default = 'dataset.tfrecord', help = 'path to tfrecord')

def parse_function(serialized_example):
  feature = tf.io.parse_single_example(
    serialized_example,
    features = {
      'x': tf.io.FixedLenFeature((), dtype = tf.string)
    })
  x = tf.io.parse_tensor(feature['x'], out_type = tf.float32)
  x = tf.reshape(x, (12,))
  return x

def main(unused_argv):
  dataframe = openpyxl.load_workbook(FLAGS.input)
  dataframe1 = dataframe.active
  data = list()
  for row in range(0, dataframe1.max_row):
    r = list()
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
      r.append(col[row].value)
    data.append(r)
  data = np.array(data).astype(np.float32)
  data = tf.constant(np.transpose(data))
  writer = tf.io.TFRecordWriter(FLAGS.output)
  for sample in data:
    trainsample = tf.train.Example(features = tf.train.Feature(
      feature = {
        'x': tf.train.Feature(bytes_list = tf.train.BytesList(value = [tf.io.serialize_tensor(sample).numpy()]))
      }))
    writer.write(trainsample.SerializeToString())
  writer.close()

if __name__ == "__main__":
  add_options()
  app.run(main)

